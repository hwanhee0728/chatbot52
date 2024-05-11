# Streamlit에 올릴때 필요한 코드 (sql 오류 fix위해), 로컬에서는 주석처리 필요
__import__('pysqlite3')
import sys
sys.modules['sqlite3'] = sys.modules.pop('pysqlite3')

# langchain 라이브러리 Import
from langchain_community.document_loaders import PyPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain_openai import OpenAIEmbeddings
from langchain_openai import ChatOpenAI
from langchain.retrievers.multi_query import MultiQueryRetriever
from langchain.chains import RetrievalQA

# 환경변수, 이미지, 데이터조작, 임시파일, 운영체계관련, 입출력, 시간 관련 라이브러리 Import
from dotenv import load_dotenv
import streamlit as st
from PIL import Image
import pandas as pd
import tempfile
import os
import io
import time

# 한글자씩 답변하기 위한, Stream을 위한 라이브러리ㅡImport
from langchain.callbacks.streaming_stdout import StreamingStdOutCallbackHandler
from langchain.callbacks.base import BaseCallbackHandler

# 소켓프로그래밍, 날짜/사간, 엑셀파일을 위한 라이브러리 Import
import socket
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 엑셀 파일에 데이터를 기록하는 함수
def log_question_to_excel(question, ip_address, timestamp):
    filename = 'question.xlsx'
    try:
        # 엑셀 파일이 이미 존재하면 로드하고, 그렇지 않으면 새 파일 생성
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet['A1'] = 'Question'
            sheet['B1'] = 'IP Address'
            sheet['C1'] = 'Timestamp'

        # 새로운 데이터 추가
        new_row = (question, ip_address, timestamp)
        sheet.append(new_row)

        # 파일 저장
        workbook.save(filename)
    except Exception as e:
        print(f"Error logging question to Excel: {e}")

# 로컬 IP 주소를 가져오는 함수
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # 이 IP는 실제로 연결되지 않지만, 루프백 주소(127.0.0.1)를 반환하지 않도록 함
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

# 환경변수 로드
load_dotenv()
password_key = os.getenv('KEY')
admin_key = os.getenv('ADMIN')

# 엑셀 파일 다운로드를 위한 함수
def download_excel():
    filename = 'question.xlsx'
    with open(filename, "rb") as file:
        btn = st.download_button(
                label="질문 로그 다운로드",
                data=file,
                file_name=filename,
                mime="application/vnd.ms-excel"
            )

# 챗봇 메인 함수
def app():

    # 화면 6:15로 분할
    col1, col2 = st.columns([6,15])

    with col1:
        st.subheader(':robot_face: Chatbot')
        image = Image.open('robot01.png')
        image = image.resize((170, 170))
        st.image(image)

        st.write("")
        st.write(':fire: KMLA Chatbot Team')
        st.write('(구환희,전지훈,권휘우)')
        st.write("")
        st.write("")
        st.write(':star: 스벅쿠폰 드려요!')   
        st.write('우측화면 챗봇 사용후,')
        st.write('만족도조사 하신분께!')
        st.write('**기간 : 5월13월~18토')
        st.write(':cupcake:추첨 : 2만원 * 3명')
        st.write(':cupcake:선정 : 2만원 * 2명')
        st.markdown('<a href="https://survey10.streamlit.app/" target="_blank">만족도조사 Click!</a>', unsafe_allow_html=True)
        st.write("")
        st.write('[ Chroma DB 컨텐츠 ]')
        st.write(':one: homepage_v20')
        st.write(':two: schoolregulation_v15')
        st.write(':three: knowhow_v15')
        st.write(':four: namuwiki_v15')
        st.write("")

        # 엑셀 다운로드
        admin_password = st.text_input(":lock: 관리자", type="password")
        if admin_password:
            if admin_password == admin_key:
                st.success("비밀번호 확인 완료")
                # 엑셀 파일 다운로드 기능
                download_excel()
            else:
                st.error("잘못된 비밀번호입니다.")

    with col2:
        st.write("")
        st.write(':heavy_check_mark: 민사고 :gun: 패스워드 넣어주세요!')
        # 패스워드 받고, 화면 보여주기위한 텍스트 입력
        password = st.text_input(":key: 안내 받은, 7 글자를 입력해 주세요 ( k * * * * * 7 )", type="password")

        if password:
            if password == password_key:

                # 비밀번호가 맞으면 화면 보여주기
                st.success("패스워드 확인 완료!")
                st.write("")
                user_input = st.text_input(":eight_pointed_black_star:민사고에 대해 질문하고 엔터를 눌러주세요!")

                st.write(':one: homepage : 국어 선생님을 알려줄래? 학교 6월 일정을 알려줄래?')
                st.write(':two: schoolregulation : 외출외박 신청은 어떻게 해? 벌점이 3점이상?')
                st.write(':three: knowhow : 비트가 뭐야? 학교에 어떤 동아리가 있어?')
                st.write(':four: namuwiki : 치킨데이가 뭐야? 바베큐 파티를 하려는데 방법은?')

                if user_input:
                    # 질문하면 아래 내용 보여주기
                    st.write("")
                    st.write(':pencil2::pencil2::pencil2::pencil2::pencil2::pencil2::pencil2::pencil2: 답변 드립니다 :pencil2::pencil2::pencil2::pencil2::pencil2::pencil2::pencil2::pencil2:')

                    # 질문을 로깅합니다.
                    log_question_to_excel(user_input, get_local_ip(), datetime.datetime.now())

                    embeddings_model = OpenAIEmbeddings()
                    db = Chroma(persist_directory="chromadb_600", embedding_function=embeddings_model)

                    # Stream구현을(한글자씩 쓰여지는 기능) 위해 Handler 만들기
                    class StreamHandler(BaseCallbackHandler):
                        def __init__(self, container, initial_text=""):
                            self.container = container
                            self.text=initial_text
                        def on_llm_new_token(self, token: str, **kwargs) -> None:
                            self.text += token
                            self.container.markdown(self.text)

                    # 질문을 받아, 답변하는 로직, 이 프로그램의 주인공.
                    # 모델 : OpenAI의 gpt-4-turbo-2024-04-10
                    # 답변 정도 : Temperature = 0
                    # DB는 Chromadb, embedding은 OpenAIEmbeddings
                    question = user_input
                    chat_box = st.empty()
                    stream_handler = StreamHandler(chat_box)                
                    llm = ChatOpenAI(model_name="gpt-4-turbo-2024-04-09", temperature=0, streaming=True, callbacks=[stream_handler])
                    qa_chain = RetrievalQA.from_chain_type(llm,retriever=db.as_retriever())         
                    qa_chain.invoke({"query": question})

            else:
                st.error("에러")

if __name__ == "__main__":
    app()